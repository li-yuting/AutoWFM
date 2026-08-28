# -*- coding: utf-8 -*-
"""shift 输出层测试：统计表列。"""
import os
import sys

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _ROOT)
sys.path.insert(0, os.path.join(_ROOT, "shift"))

from openpyxl import Workbook  # noqa: E402

from shift_test_utils import make_config, make_schedule  # noqa: E402
from utils import D_FAMILY, REST_SHIFT, WORK_SHIFTS, Z_FAMILY  # noqa: E402
from writer import _streak_shifts, _write_stats  # noqa: E402


def test_stats_employee_columns():
    s = make_schedule([
        ("甲", "A", 1.0, "正式", ["D", "A2", "A2", "A2", "OFF"]),
        ("乙", "A", 1.0, "正式", ["Z", "Z", "A1", "A4", "OFF"]),
    ])
    wb = Workbook()
    _write_stats(wb.active, s)
    ws = wb.active
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "姓名")
    headers = [ws.cell(header_row, c).value for c in range(1, 8)]
    assert headers == ["姓名", "班组", "D/D1 均衡", "Z/Z1 均衡", "A1/A4 均衡", "休息天数", "连续双休次数"]


def test_stats_values():
    s = make_schedule([
        ("甲", "A", 1.0, "正式", ["D", "D1", "Z", "A1", "A4"]),
        ("乙", "A", 1.0, "正式", ["A2", "A2", "A3", "OFF", "OFF"]),
    ], lock_values=False)
    # 甲: D/D1=2, Z/Z1=1, A1/A4=2；乙: 全 0
    wb = Workbook()
    _write_stats(wb.active, s)
    ws = wb.active
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "姓名")
    row_jia = [ws.cell(header_row + 1, c).value for c in range(3, 6)]
    row_yi = [ws.cell(header_row + 2, c).value for c in range(3, 6)]
    assert row_jia == [2, 1, 2]
    assert row_yi == [0, 0, 0]


def test_streak_shifts_family():
    assert _streak_shifts("08") == D_FAMILY
    assert _streak_shifts("18") == Z_FAMILY
    assert _streak_shifts("05") == {REST_SHIFT}
    assert _streak_shifts("04") == WORK_SHIFTS


if __name__ == "__main__":
    test_stats_employee_columns()
    test_stats_values()
    test_streak_shifts_family()
    print("test_shift_writer OK")
