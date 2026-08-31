import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import datetime as dt
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from peakflow import config, report
from tests.helpers import make_history

# Use workspace-local temp to avoid sandbox restrictions on system temp
_WS_TMP = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".test_tmp")


def _future_dates(hist, n=10):
    last = hist["date"].max()
    return [last + dt.timedelta(days=i) for i in range(1, n + 1)]


def _sample():
    hist = make_history(35)
    fd = _future_dates(hist, 10)
    from peakflow.forecast import three_band_forecast, backtest_sigma
    df, _ = three_band_forecast(hist, fd)
    sig = {"在线": backtest_sigma(hist), "热线": backtest_sigma(hist)}
    return df, sig


def test_write_report():
    os.makedirs(_WS_TMP, exist_ok=True)
    try:
        df, sig = _sample()
        meta = {"生成时间": "2026-08-13 12:00:00", "预测天数": 10}
        out = Path(_WS_TMP) / "out.xlsx"
        report.write_report(df, df, sig, meta, out)
        wb = load_workbook(out)
        assert wb.sheetnames == ["总览", "在线-分类型明细", "热线-分类型明细", "回测误差", "运行说明"]
        ws = wb["总览"]
        assert ws.max_row > 10
        ws2 = wb["运行说明"]
        assert "预测天数" in str([c.value for c in ws2["A"]])
    finally:
        shutil.rmtree(_WS_TMP, ignore_errors=True)
    print("PASS test_write_report")


def test_report_has_totals_in_detail():
    os.makedirs(_WS_TMP, exist_ok=True)
    try:
        df, sig = _sample()
        out = Path(_WS_TMP) / "out.xlsx"
        report.write_report(df, df, sig, {}, out)
        wb = load_workbook(out)
        ws = wb["在线-分类型明细"]
        types = {ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)}
        assert "合计" in types
    finally:
        shutil.rmtree(_WS_TMP, ignore_errors=True)
    print("PASS test_report_has_totals_in_detail")


def main():
    test_write_report()
    test_report_has_totals_in_detail()
    print("\nAll tests passed!")


if __name__ == "__main__":
    main()
