# -*- coding: utf-8 -*-
"""shift 全流程冒烟：在线模板注入 B/C 需求 → 排班 → 校验无 ERROR。"""
import os
import shutil
import sys
import time

from openpyxl import load_workbook

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _ROOT)
sys.path.insert(0, os.path.join(_ROOT, "shift"))

from reader import read_schedule  # noqa: E402
from scheduler import SchedulerConfig, run_scheduler  # noqa: E402
from validators import validate_schedule  # noqa: E402
from writer import write_schedule  # noqa: E402

_TMP = os.path.join(_ROOT, "tests", ".test_tmp")


def test_pipeline_smoke():
    src = os.path.join(_ROOT, "shift", "排班计划 - 在线.xlsx")
    if not os.path.isfile(src):
        print("test_shift_pipeline SKIP: 模板文件不入库, 本地缺失")
        return
    os.makedirs(_TMP, exist_ok=True)
    stamp = time.time_ns()
    work = os.path.join(_TMP, f"smoke_{stamp}.xlsx")
    out = os.path.join(_TMP, f"smoke_out_{stamp}.xlsx")
    shutil.copyfile(src, work)

    wb = load_workbook(work)
    ws = wb["需求"]
    for row in range(2, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip().upper()
        if label in ("B", "C"):
            for col in range(2, ws.max_column + 1):
                ws.cell(row, col, 2)
    wb.save(work)

    schedule = read_schedule(work)
    config = SchedulerConfig(preset_rest_days=8, z_min_consecutive=2, z_max_consecutive=3)
    run_scheduler(schedule, config)
    schedule.warnings.clear()
    validate_schedule(schedule, config)
    write_schedule(schedule, out)

    errors = [w for w in schedule.warnings if w.severity == "ERROR"]
    assert not errors, [f"{w.check_id}{w.employee}{w.date}: {w.message}" for w in errors[:10]]

    bases = [c.base_shift for e in schedule.employees for c in e.schedule]
    assert "B" in bases, "B 需求>0 应有排班产出"
    assert os.path.isfile(out)


if __name__ == "__main__":
    test_pipeline_smoke()
    print("test_shift_pipeline OK")
