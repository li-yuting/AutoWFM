from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from models import DailyDemand, Employee, Schedule, ShiftCell
from utils import SHIFT_ORDER, clean_text, number

DEMAND_SHEET = "需求"
SCHEDULE_SHEET = "班表"
META_COLUMNS = 4


def read_schedule(path: str | Path, history_days: int = 6) -> Schedule:
    workbook_path = str(Path(path))
    wb = load_workbook(workbook_path, data_only=False)
    demand_ws = wb[DEMAND_SHEET] if DEMAND_SHEET in wb.sheetnames else wb.worksheets[0]
    schedule_ws = wb[SCHEDULE_SHEET] if SCHEDULE_SHEET in wb.sheetnames else wb.worksheets[1]

    schedule_dates = [
        schedule_ws.cell(1, col).value
        for col in range(META_COLUMNS + 1, schedule_ws.max_column + 1)
    ]
    date_columns = list(range(META_COLUMNS + 1, schedule_ws.max_column + 1))
    demands = _read_demands(demand_ws, schedule_dates)

    employees: list[Employee] = []
    for row in range(2, schedule_ws.max_row + 1):
        name = clean_text(schedule_ws.cell(row, 4).value)
        if not name:
            continue
        person_type = clean_text(schedule_ws.cell(row, 1).value)
        coefficient = number(schedule_ws.cell(row, 2).value, default=1.0)
        group = clean_text(schedule_ws.cell(row, 3).value)
        cells: list[ShiftCell] = []
        for idx, col in enumerate(date_columns):
            value = schedule_ws.cell(row, col).value
            locked = value is not None and str(value).strip() != ""
            cells.append(
                ShiftCell(
                    value=value,
                    original_value=value,
                    is_locked=locked,
                    is_historical=idx < history_days,
                    row=row,
                    column=col,
                )
            )
        employees.append(
            Employee(
                name=name,
                group=group,
                coefficient=coefficient,
                is_phase3="三期" in person_type,
                schedule=cells,
                row_index=row,
                person_type=person_type,
            )
        )

    return Schedule(
        employees=employees,
        dates=schedule_dates,
        demands=demands,
        workbook_path=workbook_path,
        history_days=history_days,
        schedule_sheet_name=schedule_ws.title,
        demand_sheet_name=demand_ws.title,
        date_columns=date_columns,
    )


def _read_demands(ws, schedule_dates: list) -> list[DailyDemand]:
    demand_by_date: dict[object, dict[str, float]] = {}
    for col in range(2, ws.max_column + 1):
        demand_by_date[ws.cell(1, col).value] = {}

    for row in range(2, ws.max_row + 1):
        shift = clean_text(ws.cell(row, 1).value).upper()
        if not shift:
            continue
        for col in range(2, ws.max_column + 1):
            date = ws.cell(1, col).value
            demand_by_date.setdefault(date, {})[shift] = number(ws.cell(row, col).value)

    demands: list[DailyDemand] = []
    for date in schedule_dates:
        row = {shift: 0.0 for shift in SHIFT_ORDER + ("OFF",)}
        row.update(demand_by_date.get(date, {}))
        demands.append(DailyDemand(date=date, demand=row))
    return demands
