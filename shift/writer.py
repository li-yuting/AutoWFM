from __future__ import annotations

from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models import Schedule
from utils import (
    A_BALANCE_SHIFTS,
    D_BALANCE_SHIFTS,
    D_FAMILY,
    REST_SHIFT,
    SHIFT_ORDER,
    WORK_SHIFTS,
    Z_BALANCE_SHIFTS,
    Z_FAMILY,
    date_label,
)


def write_schedule(schedule: Schedule, output_path: str | Path) -> None:
    wb = load_workbook(schedule.workbook_path)
    ws = wb[schedule.schedule_sheet_name]

    for employee in schedule.employees:
        for cell in employee.schedule:
            ws.cell(employee.row_index, cell.column).value = cell.value

    _highlight_schedule(ws, schedule)

    if "统计" in wb.sheetnames:
        del wb["统计"]
    stats = wb.create_sheet("统计")
    _write_stats(stats, schedule)
    wb.save(output_path)


# 级别 -> 单元格填充色（浅色，不遮文字）；同格多级命中时取 ERROR > WARN > INFO
# 配色方案 B：Excel 经典条件格式（浅红 / 黄 / 蓝）
_SEVERITY_FILL = {
    "ERROR": PatternFill("solid", fgColor="FFC7CE"),
    "WARN": PatternFill("solid", fgColor="FFEB9C"),
    "INFO": PatternFill("solid", fgColor="BDD7EE"),
}
_SEVERITY_ORDER = {"ERROR": 0, "WARN": 1, "INFO": 2}


def _highlight_schedule(ws, schedule: Schedule) -> None:
    """按验证警告为班表 sheet 的班次单元格填充高亮色。"""
    date_cols = {d: col for d, col in zip(schedule.dates, schedule.date_columns)}
    cells: dict[tuple[int, int], str] = {}
    for w in schedule.warnings:
        if not w.employee or w.date is None or w.severity not in _SEVERITY_FILL:
            continue
        emp = next((e for e in schedule.employees if e.name == w.employee), None)
        if emp is None or w.date not in date_cols:
            continue
        idx = schedule.dates.index(w.date)
        col = date_cols[w.date]
        if w.check_id in ("04", "05", "08", "18"):
            start, end = _containing_streak(emp, idx, _streak_shifts(w.check_id))
        elif w.check_id == "14":
            start, end = _rest_gap_range(emp, idx)
        else:
            start, end = idx, idx
        for i in range(start, end + 1):
            key = (emp.row_index, date_cols[schedule.dates[i]])
            if key not in cells or _SEVERITY_ORDER[w.severity] < _SEVERITY_ORDER[cells[key]]:
                cells[key] = w.severity
    for (row, col), sev in cells.items():
        ws.cell(row, col).fill = _SEVERITY_FILL[sev]


def _streak_shifts(check_id: str) -> set[str]:
    if check_id == "05":
        return {REST_SHIFT}
    if check_id == "08":
        return D_FAMILY
    if check_id == "18":
        return Z_FAMILY
    return WORK_SHIFTS


def _streaks(employee, shifts: set[str]) -> list[tuple[int, int]]:
    result = []
    start = None
    for idx, cell in enumerate(employee.schedule):
        if cell.base_shift in shifts:
            if start is None:
                start = idx
        elif start is not None:
            result.append((start, idx - 1))
            start = None
    if start is not None:
        result.append((start, len(employee.schedule) - 1))
    return result


def _containing_streak(employee, idx: int, shifts: set[str]) -> tuple[int, int]:
    for start, end in _streaks(employee, shifts):
        if start <= idx <= end:
            return start, end
    return idx, idx


def _rest_blocks(employee) -> list[tuple[int, int]]:
    blocks = []
    start = None
    for idx, cell in enumerate(employee.schedule):
        if cell.base_shift == REST_SHIFT:
            if start is None:
                start = idx
        elif start is not None:
            blocks.append((start, idx - 1))
            start = None
    if start is not None:
        blocks.append((start, len(employee.schedule) - 1))
    return blocks


def _rest_gap_range(employee, idx: int) -> tuple[int, int]:
    """休息间隔警告(14)：高亮 前一休息块 → 上班段 → 本次休息块 的整段。"""
    blocks = _rest_blocks(employee)
    pos = next((i for i, (s, e) in enumerate(blocks) if s <= idx <= e), None)
    if pos is None:
        return idx, idx
    prev_start = blocks[pos - 1][0] if pos > 0 else idx
    return prev_start, blocks[pos][1]


def _write_stats(ws, schedule: Schedule) -> None:
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    normal_font = Font(name="Microsoft YaHei")

    row = 1
    row = _section(ws, row, "每日满足情况", header_fill, header_font)
    ws.append(["日期", "班次", "需求", "实际", "差异"])
    _style_header(ws[row])
    row += 1
    for day_index, demand in enumerate(schedule.adjusted_demands or []):
        actual = _actual_by_shift(schedule, day_index)
        for shift in SHIFT_ORDER + ("OFF",):
            target = demand.get(shift)
            value = actual.get(shift, 0.0)
            ws.append([date_label(schedule.dates[day_index]), shift, round(target, 2), round(value, 2), round(value - target, 2)])
            row += 1

    row += 2
    row = _section(ws, row, "员工统计", header_fill, header_font)
    ws.append(["姓名", "班组", "D/D1 均衡", "Z/Z1 均衡", "A1/A4 均衡", "休息天数", "连续双休次数"])
    _style_header(ws[row])
    row += 1
    for employee in schedule.employees:
        d_count = 0
        z_count = 0
        a_count = 0
        rest_days = 0
        double_rests = 0
        prev_rest = False
        for idx in schedule.active_indexes:
            base = employee.schedule[idx].base_shift
            if base in D_BALANCE_SHIFTS:
                d_count += 1
            if base in Z_BALANCE_SHIFTS:
                z_count += 1
            if base in A_BALANCE_SHIFTS:
                a_count += 1
            is_rest = base == "OFF"
            if is_rest:
                rest_days += 1
            if is_rest and prev_rest:
                double_rests += 1
            prev_rest = is_rest
        ws.append([employee.name, employee.group, d_count, z_count, a_count, rest_days, double_rests])
        row += 1

    row += 2
    row = _section(ws, row, "OFF/A3 调整", header_fill, header_font)
    ws.append(["日期", "OFF转A3", "A3转OFF"])
    _style_header(ws[row])
    row += 1
    for demand in schedule.adjusted_demands:
        if demand.off_to_a3 or demand.a3_to_off:
            ws.append([date_label(demand.date), round(demand.off_to_a3, 2), round(demand.a3_to_off, 2)])
            row += 1

    row += 2
    row = _section(ws, row, "警告信息", header_fill, header_font)
    ws.append(["编号", "级别", "员工", "日期", "描述"])
    _style_header(ws[row])
    row += 1
    for warning in schedule.warnings:
        ws.append([warning.check_id, warning.severity, warning.employee, date_label(warning.date), warning.message])
        row += 1

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 18
    for cells in ws.iter_rows():
        for cell in cells:
            if cell.font:
                font = copy(cell.font)
                font.name = "Microsoft YaHei"
                cell.font = font
            else:
                cell.font = normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _actual_by_shift(schedule: Schedule, day_index: int) -> dict[str, float]:
    totals = defaultdict(float)
    for employee in schedule.employees:
        base = employee.schedule[day_index].base_shift
        if base:
            totals[base] += employee.coefficient
    return dict(totals)


def _section(ws, row: int, title: str, fill, font) -> int:
    ws.cell(row, 1).value = title
    ws.cell(row, 1).fill = fill
    ws.cell(row, 1).font = font
    return row + 1


def _style_header(cells) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in cells:
        cell.fill = fill
        cell.font = Font(bold=True, name="Microsoft YaHei")
